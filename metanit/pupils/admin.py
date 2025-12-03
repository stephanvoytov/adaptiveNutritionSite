import urllib

from django.shortcuts import render
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.contrib import admin
from openpyxl.utils import get_column_letter

from .models import Class, Pupil, Dish, DailyMenu, WeeklyBreakfasts


@admin.register(Class)
class ClassAdmin(admin.ModelAdmin):
    list_display = ['name', 'number_of_pupils']
    list_filter = ['name']


@admin.register(Pupil)
class PupilAdmin(admin.ModelAdmin):
    list_display = ['last_name', 'first_name', 'class_group']
    list_filter = ['class_group']
    search_fields = ['last_name', 'first_name']


@admin.register(Dish)
class DishAdmin(admin.ModelAdmin):
    list_display = ['name', 'short_name']
    search_fields = ['name', 'short_name']


@admin.register(DailyMenu)
class DailyMenuAdmin(admin.ModelAdmin):
    list_display = ['date', 'option_1', 'option_2']
    list_filter = ['date']
    date_hierarchy = 'date'


@admin.register(WeeklyBreakfasts)
class WeeklyBreakfastAdmin(admin.ModelAdmin):
    list_display = ['pupil', 'get_class_group', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday',
                    'week_start_date', 'choices_count']
    list_filter = ['pupil__class_group', 'week_start_date']
    search_fields = ['pupil__first_name', 'pupil__last_name']

    def get_class_group(self, obj):
        return obj.pupil.class_group

    get_class_group.short_description = 'Класс'

    def choices_count(self, obj):
        count = sum([1 for field in [obj.monday, obj.tuesday, obj.wednesday,
                                     obj.thursday, obj.friday] if field])
        return f"{count}/5"

    choices_count.short_description = 'Выбрано'

    def get_urls(self):
        urls = super().get_urls()
        from django.urls import path
        custom_urls = [
            path('statistics/', self.admin_site.admin_view(self.statistics_view),
                 name='pupils_weeklybreakfasts_statistics'),
            path('export-excel/', self.admin_site.admin_view(self.export_to_excel),
                 name='pupils_weeklybreakfasts_export_excel'),
        ]
        return custom_urls + urls

    def statistics_view(self, request):
        """Кастомная страница статистики"""
        from datetime import datetime, timedelta

        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())

        stats_data = get_weekly_statistics(week_start)

        context = {
            'title': f'Статистика завтраков на неделю с {week_start} по {week_start + timedelta(days=4)}',
            'classes_stats': stats_data['classes_stats'],
            'week_start': week_start,
        }

        return render(request, 'admin/breakfast_statistics.html', context)

    def changelist_view(self, request, extra_context=None):
        """Добавляем ссылку на статистику в список"""
        extra_context = extra_context or {}
        extra_context['show_statistics_link'] = True
        return super().changelist_view(request, extra_context=extra_context)

    def export_to_excel(self, request):
        """Экспорт с листами статистики и листами учеников"""
        from .models import Class, Pupil, WeeklyBreakfasts
        from datetime import datetime, timedelta

        from openpyxl.styles import PatternFill

        not_chosen_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        chosen_style = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())

        # Создаем Excel книгу
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Стили
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center')

        # 1.\
        stats_data = get_weekly_statistics(week_start)
        ws_stats = wb.create_sheet(title="Статистика")

        # Заголовок статистики
        ws_stats.merge_cells('A1:E1')
        ws_stats['A1'] = f'Статистика завтраков с {week_start} по {stats_data["week_end"]}'
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats['A1'].alignment = center_align

        row = 3

        for class_stat in stats_data['classes_stats']:
            class_obj = class_stat['class']

            # Заголовок класса
            ws_stats.merge_cells(f'A{row}:E{row}')
            ws_stats[f'A{row}'] = f'Класс: {class_obj.name}'
            ws_stats[f'A{row}'].font = bold_font
            row += 1

            # Всего учеников
            ws_stats[f'A{row}'] = f'Всего учеников: {class_stat["total_pupils"]}'
            row += 1

            # Заголовки таблицы
            headers = ['День недели', 'Вариант 1', 'Вариант 2', 'Не выбрано', 'Выбрали']
            for col, header in enumerate(headers, 1):
                cell = ws_stats.cell(row=row, column=col, value=header)
                cell.font = bold_font
                cell.fill = header_fill

            row += 1

            # Статистика по дням
            days_mapping = {
                'monday': 'Понедельник',
                'tuesday': 'Вторник',
                'wednesday': 'Среда',
                'thursday': 'Четверг',
                'friday': 'Пятница'
            }

            for day_field, day_name in days_mapping.items():
                day_data = class_stat['day_stats'][day_field]
                dishes_list = list(day_data['dishes'].items())

                option1 = f"{dishes_list[0][0]}: {dishes_list[0][1]}" if dishes_list else "-"
                option2 = f"{dishes_list[1][0]}: {dishes_list[1][1]}" if len(dishes_list) > 1 else "-"

                # Записываем строку
                ws_stats.cell(row=row, column=1, value=day_name)
                ws_stats.cell(row=row, column=2, value=option1)
                ws_stats.cell(row=row, column=3, value=option2)
                ws_stats.cell(row=row, column=4, value=day_data['not_chosen'])
                ws_stats.cell(row=row, column=5, value=day_data['chosen'])

                row += 1

            # Пустая строка между классами
            row += 2

        # 2. ЛИСТЫ С УЧЕНИКАМИ
        for class_obj in Class.objects.all().order_by('name'):
            ws_pupils = wb.create_sheet(title=f"Ученики {class_obj.name}")

            # Заголовок
            ws_pupils.merge_cells('A1:H1')
            ws_pupils['A1'] = f'Выборы завтраков - {class_obj.name} (неделя {week_start})'
            ws_pupils['A1'].font = Font(bold=True, size=14)
            ws_pupils['A1'].alignment = center_align

            # Заголовки таблицы
            headers = ['Фамилия', 'Имя', 'Класс', 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']
            for col, header in enumerate(headers, 1):
                cell = ws_pupils.cell(row=3, column=col, value=header)
                cell.font = bold_font
                cell.fill = header_fill

            # Данные учеников
            pupils = Pupil.objects.filter(class_group=class_obj).order_by('last_name', 'first_name')
            row_pupil = 4

            for pupil in pupils:
                try:
                    breakfast = WeeklyBreakfasts.objects.get(pupil=pupil, week_start_date=week_start)
                except WeeklyBreakfasts.DoesNotExist:
                    breakfast = None

                # Данные ученика
                ws_pupils.cell(row=row_pupil, column=1, value=pupil.last_name)
                ws_pupils.cell(row=row_pupil, column=2, value=pupil.first_name)
                ws_pupils.cell(row=row_pupil, column=3, value=class_obj.name)

                days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
                for col, day_field in enumerate(days, 4):
                    cell = ws_pupils.cell(row=row_pupil, column=col)
                    if breakfast and getattr(breakfast, day_field):
                        cell.fill = chosen_style
                        cell.value = getattr(breakfast, day_field).short_name
                    else:
                        cell.fill = not_chosen_fill
                        cell.value = "Не выбрано"

                row_pupil += 1

            # Настраиваем ширину колонок
            column_widths = [15, 15, 10, 20, 20, 20, 20, 20]
            for i, width in enumerate(column_widths, 1):
                ws_pupils.column_dimensions[get_column_letter(i)].width = width

        # Настраиваем ширину колонок для листа статистики
        stats_widths = [15, 25, 25, 12, 12]
        for i, width in enumerate(stats_widths, 1):
            ws_stats.column_dimensions[get_column_letter(i)].width = width

        # Создаем HTTP ответ с файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'завтраки_{week_start}.xlsx'
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"

        wb.save(response)
        return response


def get_weekly_statistics(week_start=None):
    """Общая функция для получения статистики за неделю"""
    from django.db.models import Count
    from datetime import datetime, timedelta

    if week_start is None:
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())

    from .models import Class, WeeklyBreakfasts

    classes_stats = []

    for class_obj in Class.objects.all():
        total_pupils = class_obj.number_of_pupils
        day_stats = {}
        days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']

        for day in days:
            choices = WeeklyBreakfasts.objects.filter(
                week_start_date=week_start,
                pupil__class_group=class_obj
            ).exclude(**{f"{day}__isnull": True})

            dish_counts = choices.values(
                f'{day}__short_name'
            ).annotate(count=Count('id')).order_by('-count')

            dish_stats = {}
            for item in dish_counts:
                dish_stats[item[f'{day}__short_name']] = item['count']

            day_stats[day] = {
                'dishes': dish_stats,
                'not_chosen': total_pupils - choices.count(),
                'chosen': choices.count()
            }

        classes_stats.append({
            'class': class_obj,
            'total_pupils': total_pupils,
            'day_stats': day_stats
        })

    return {
        'classes_stats': classes_stats,
        'week_start': week_start,
        'week_end': week_start + timedelta(days=4)
    }
